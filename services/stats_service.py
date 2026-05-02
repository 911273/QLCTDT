# services/stats_service.py
"""
StatsService — Phân tích thống kê, kiểm định chất lượng chương trình đào tạo.
Xuất báo cáo Excel (ABET/AUN style) và thống kê nội bộ.
Dùng openpyxl cho Excel, không cần pandas.
"""

import os
from datetime import datetime
from typing import List, Optional, Dict, Any


class StatsService:
    """Cung cấp phân tích thống kê và xuất báo cáo Excel."""

    def __init__(self, db):
        self.db = db

    # ── Thống kê Dashboard ───────────────────────────────────────────────────
    def get_overall_dashboard_stats(self) -> dict:
        """Thống kê tổng quan cho toàn hệ thống (Dashboard v2.0)."""
        db = self.db
        
        # 1. Cơ bản
        total_hp = db.conn.execute("SELECT COUNT(*) FROM hoc_phan").fetchone()[0] or 0
        digital_hp = total_hp # Tạm thời coi 100% là số hóa nếu có trong DB hnay
        
        # 2. Đã validate (có kế hoạch đánh giá đủ 100%)
        validated_rows = db.conn.execute("""
            SELECT hp_id, SUM(ty_trong) as s 
            FROM ke_hoach_kiem_tra GROUP BY hp_id HAVING s >= 99.9
        """).fetchall()
        validated_hp = len(validated_rows)
        
        # 3. Danh sách cập nhật lâu nhất (Outdated)
        outdated = db.conn.execute("""
            SELECT ma, ten_viet, COALESCE(ngay_cap_nhat, ngay_tao, 'N/A') 
            FROM hoc_phan ORDER BY ngay_cap_nhat ASC, ngay_tao ASC LIMIT 5
        """).fetchall()
        
        # 4. PLO ít được hỗ trợ nhất (Weak PLOs)
        weak_plos = db.conn.execute("""
            SELECT cdr_ma, COUNT(*) as cnt 
            FROM clo 
            WHERE cdr_ma IS NOT NULL AND cdr_ma != ''
            GROUP BY cdr_ma ORDER BY cnt ASC LIMIT 5
        """).fetchall()
        
        return {
            'total_hp': total_hp,
            'digital_hp': digital_hp,
            'validated_hp': validated_hp,
            'outdated': outdated,
            'weak_plos': weak_plos,
            'timestamp': datetime.now().strftime('%d/%m/%Y %H:%M')
        }

    # ── Thống kê HP ──────────────────────────────────────────────────────────

    def get_hp_summary(self, hp_id: int) -> dict:
        """Thống kê tổng quan 1 học phần."""
        hp = dict(self.db.get_hoc_phan(hp_id) or {})
        clos = [dict(r) for r in self.db.get_clo(hp_id)]
        mts = [dict(r) for r in self.db.get_muc_tieu(hp_id)]
        nd_lt = [dict(r) for r in self.db.get_noi_dung(hp_id, 'lt')]
        nd_th = [dict(r) for r in self.db.get_noi_dung(hp_id, 'th')]
        kt = [dict(r) for r in self.db.get_ke_hoach_kt(hp_id)]
        hl = [dict(r) for r in self.db.get_hoc_lieu(hp_id)]

        total_lt = sum(float(r.get('gio_lt') or 0) for r in nd_lt)
        total_bt = sum(float(r.get('gio_bt') or 0) for r in nd_lt)
        total_th = sum(float(r.get('gio_th_tn') or 0) for r in nd_lt + nd_th)
        total_kt_weight = sum(float(r.get('ty_trong') or 0) for r in kt)

        # Phân bổ CLO theo PLO
        plo_map = {}
        for clo in clos:
            plo = clo.get('cdr_ma', 'N/A') or 'N/A'
            plo_map.setdefault(plo, []).append(clo.get('ma', ''))

        return {
            'hp': hp,
            'clo_count': len([c for c in clos if not c.get('la_tieu_de_nhom')]),
            'mt_count': len([m for m in mts if not m.get('la_tieu_de_nhom')]),
            'nd_lt_count': len(nd_lt),
            'nd_th_count': len(nd_th),
            'kt_count': len(kt),
            'hl_chinh': len([h for h in hl if h.get('loai') == 'chinh']),
            'hl_tham_khao': len([h for h in hl if h.get('loai') == 'tham_khao']),
            'total_lt_hours': total_lt,
            'total_bt_hours': total_bt,
            'total_th_hours': total_th,
            'total_kt_weight': total_kt_weight,
            'plo_distribution': {k: len(v) for k, v in plo_map.items()},
            'clo_levels': {
                'I': len([c for c in clos if c.get('level_irm') == 'I']),
                'R': len([c for c in clos if c.get('level_irm') == 'R']),
                'M': len([c for c in clos if c.get('level_irm') == 'M']),
            }
        }

    def get_clo_plo_matrix(self, hp_id: int) -> List[list]:
        """
        Tạo ma trận CLO × PLO dạng list of lists.
        Row 0: header [Tên HP, PLO1, PLO2, ...]
        Row 1+: [CLO code, I/R/M hoặc '']
        """
        clos = [dict(r) for r in self.db.get_clo(hp_id)
                if not dict(r).get('la_tieu_de_nhom')]

        # Lấy danh sách PLO từ DB (nếu có CTĐT liên kết), hoặc từ clo.cdr_ma
        plos_in_clos = sorted(set(
            c.get('cdr_ma', '') for c in clos if c.get('cdr_ma')
        ))

        if not plos_in_clos:
            return [['CLO', 'PLO'], *[[c.get('ma'), c.get('level_irm', 'I')] for c in clos]]

        header = ['CLO / PLO'] + plos_in_clos
        rows = [header]
        for clo in clos:
            row = [clo.get('ma', '')]
            clo_plo = clo.get('cdr_ma', '')
            clo_lvl = clo.get('level_irm', 'I')
            for plo in plos_in_clos:
                row.append(clo_lvl if plo == clo_plo else '')
            rows.append(row)
        return rows

    # ── Export Excel ──────────────────────────────────────────────────────────

    def export_hp_report_excel(self, hp_id: int, output_path: str) -> bool:
        """
        Xuất báo cáo Excel tổng hợp cho 1 học phần:
        - Sheet 1: Thông tin chung + thống kê
        - Sheet 2: Ma trận CLO-PLO
        - Sheet 3: Kế hoạch đánh giá
        """
        try:
            import openpyxl
            from openpyxl.styles import (
                PatternFill, Font, Alignment, Border, Side, GradientFill
            )
        except ImportError:
            raise RuntimeError("Thư viện 'openpyxl' chưa cài. Chạy: pip install openpyxl")

        wb = openpyxl.Workbook()

        # Colors
        HDR_FILL  = PatternFill("solid", fgColor="1e3a5f")
        HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
        SUB_FILL  = PatternFill("solid", fgColor="2d6a9f")
        SUB_FONT  = Font(bold=True, color="FFFFFF", size=9)
        ALT_FILL  = PatternFill("solid", fgColor="e8f4fb")
        CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT      = Alignment(horizontal="left", vertical="center", wrap_text=True)
        THIN      = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        summary = self.get_hp_summary(hp_id)
        hp = summary['hp']

        # ── Sheet 1: Thông tin chung ─────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Thông tin chung"
        ws1.column_dimensions['A'].width = 28
        ws1.column_dimensions['B'].width = 40

        def _write_hdr(cell, txt):
            ws1[cell] = txt
            ws1[cell].fill = HDR_FILL
            ws1[cell].font = HDR_FONT
            ws1[cell].alignment = CENTER
            ws1[cell].border = THIN

        def _write_lbl(row, label, val):
            a = f"A{row}"
            b = f"B{row}"
            ws1[a] = label
            ws1[a].fill = ALT_FILL if row % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            ws1[a].alignment = LEFT
            ws1[a].border = THIN
            ws1[a].font = Font(bold=True, size=9)
            ws1[b] = val
            ws1[b].alignment = LEFT
            ws1[b].border = THIN
            ws1[b].font = Font(size=9)

        ws1.merge_cells('A1:B1')
        ws1['A1'] = f"BÁO CÁO HỌC PHẦN — {hp.get('ten_viet', '')}"
        ws1['A1'].fill = HDR_FILL
        ws1['A1'].font = Font(bold=True, color="FFFFFF", size=13)
        ws1['A1'].alignment = CENTER
        ws1.row_dimensions[1].height = 30
        ws1.merge_cells('A2:B2')
        ws1['A2'] = f"Xuất ngày: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws1['A2'].alignment = CENTER
        ws1['A2'].font = Font(italic=True, size=9, color="666666")

        rows_data = [
            ("Mã học phần", hp.get('ma', '')),
            ("Tên tiếng Việt", hp.get('ten_viet', '')),
            ("Tên tiếng Anh", hp.get('ten_anh', '')),
            ("Số tín chỉ", hp.get('so_tin_chi', 0)),
            ("Loại", hp.get('loai', '')),
            ("Tính chất", hp.get('tinh_chat', '')),
            ("Giờ Lý thuyết", f"{summary['total_lt_hours']:.0f}h / kế hoạch: {hp.get('gio_lt', 0)}h"),
            ("Giờ Bài tập", f"{summary['total_bt_hours']:.0f}h / kế hoạch: {hp.get('gio_bt', 0)}h"),
            ("Giờ TH/TN", f"{summary['total_th_hours']:.0f}h / kế hoạch: {hp.get('gio_th_tn', 0)}h"),
            ("Số CLO", summary['clo_count']),
            ("Số mục tiêu", summary['mt_count']),
            ("Số nội dung LT", summary['nd_lt_count']),
            ("Số nội dung TH", summary['nd_th_count']),
            ("Số bài đánh giá", summary['kt_count']),
            ("Tổng trọng số ĐG", f"{summary['total_kt_weight']:.0f}%"),
            ("Tài liệu chính", summary['hl_chinh']),
            ("Tài liệu tham khảo", summary['hl_tham_khao']),
            ("Trạng thái", hp.get('trang_thai', 'nhap')),
        ]
        for i, (lbl, val) in enumerate(rows_data, start=3):
            _write_lbl(i, lbl, val)

        # Phân bổ CLO theo Level
        r = len(rows_data) + 4
        ws1.merge_cells(f'A{r}:B{r}')
        ws1[f'A{r}'] = "Phân bổ CLO theo mức độ I/R/M"
        ws1[f'A{r}'].fill = SUB_FILL
        ws1[f'A{r}'].font = SUB_FONT
        ws1[f'A{r}'].alignment = CENTER
        r += 1
        for lvl, cnt in summary['clo_levels'].items():
            _write_lbl(r, f"Mức {lvl}", cnt)
            r += 1

        # ── Sheet 2: Ma trận CLO-PLO ─────────────────────────────────────────
        ws2 = wb.create_sheet("Ma trận CLO-PLO")
        matrix = self.get_clo_plo_matrix(hp_id)

        IRM_COLORS = {
            'I': PatternFill("solid", fgColor="a8d8a8"),   # Xanh nhạt
            'R': PatternFill("solid", fgColor="ffd966"),   # Vàng
            'M': PatternFill("solid", fgColor="f4b183"),   # Cam
        }

        for r_idx, row in enumerate(matrix):
            for c_idx, val in enumerate(row):
                cell = ws2.cell(row=r_idx+1, column=c_idx+1, value=val)
                cell.alignment = CENTER
                cell.border = THIN
                if r_idx == 0:
                    cell.fill = HDR_FILL
                    cell.font = HDR_FONT
                elif c_idx == 0:
                    cell.font = Font(bold=True, size=9)
                    cell.fill = ALT_FILL
                elif val in IRM_COLORS:
                    cell.fill = IRM_COLORS[val]
                    cell.font = Font(bold=True, size=9)

        for i in range(1, len(matrix[0]) + 1 if matrix else 2):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 10
        ws2.column_dimensions['A'].width = 14

        # ── Sheet 3: Kế hoạch đánh giá ───────────────────────────────────────
        ws3 = wb.create_sheet("Kế hoạch đánh giá")
        kt_rows = [dict(r) for r in self.db.get_ke_hoach_kt(hp_id)]

        headers3 = ['#', 'Nhóm', 'Nội dung', 'Hình thức', 'Thời gian',
                    'Thang điểm', 'Trọng số (%)', 'CLO liên quan']
        ws3.column_dimensions['A'].width = 5
        ws3.column_dimensions['B'].width = 14
        ws3.column_dimensions['C'].width = 28
        ws3.column_dimensions['D'].width = 18
        ws3.column_dimensions['E'].width = 14
        ws3.column_dimensions['F'].width = 12
        ws3.column_dimensions['G'].width = 12
        ws3.column_dimensions['H'].width = 20

        for cidx, h in enumerate(headers3, 1):
            c = ws3.cell(row=1, column=cidx, value=h)
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = CENTER
            c.border = THIN

        for i, kt in enumerate(kt_rows, 1):
            row_data = [
                i,
                ('Thường xuyên' if kt.get('nhom') == 'thuong_xuyen' else
                 'Giữa kỳ' if kt.get('nhom') == 'giua_ky' else
                 'Cuối kỳ' if kt.get('nhom') == 'cuoi_ky' else kt.get('nhom', '')),
                kt.get('noi_dung', ''),
                kt.get('hinh_thuc', ''),
                kt.get('thoi_gian', ''),
                kt.get('thang_diem', ''),
                kt.get('ty_trong', ''),
                kt.get('clo_lien_quan', ''),
            ]
            fill = ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for cidx, val in enumerate(row_data, 1):
                c = ws3.cell(row=i+1, column=cidx, value=val)
                c.fill = fill
                c.alignment = CENTER if cidx in (1, 6, 7) else LEFT
                c.border = THIN
                c.font = Font(size=9)

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        return True

    def export_ctdt_matrix_excel(self, ctdt_id: int, output_path: str) -> bool:
        """
        Xuất ma trận CLO-PLO toàn bộ CTĐT sang Excel.
        Mỗi hàng = 1 học phần, cột = PLO, ô = mức độ I/R/M cao nhất.
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        except ImportError:
            raise RuntimeError("Thư viện 'openpyxl' chưa cài. Chạy: pip install openpyxl")

        # Lấy danh sách HP trong CTĐT
        hp_rows = self.db.conn.execute("""
            SELECT chp.hp_id, hp.ma, hp.ten_viet, hp.so_tin_chi, chp.khoi_kien_thuc
            FROM ctdt_hoc_phan chp
            JOIN hoc_phan hp ON hp.id = chp.hp_id
            WHERE chp.ctdt_id=?
            ORDER BY chp.thu_tu, hp.ten_viet
        """, (ctdt_id,)).fetchall()

        # Lấy danh sách PLO
        plo_rows = self.db.conn.execute(
            "SELECT ma, mo_ta FROM ctdt_plo WHERE ctdt_id=? ORDER BY ma",
            (ctdt_id,)
        ).fetchall()

        if not hp_rows or not plo_rows:
            raise ValueError("CTĐT không có HP hoặc PLO. Vui lòng kiểm tra dữ liệu.")

        plos = [dict(r) for r in plo_rows]
        plo_mas = [p['ma'] for p in plos]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CLO-PLO Matrix"

        HDR_FILL  = PatternFill("solid", fgColor="1e3a5f")
        HDR_FONT  = Font(bold=True, color="FFFFFF", size=9)
        ALT_FILL  = PatternFill("solid", fgColor="e8f4fb")
        ITM_FILL  = {
            'I': PatternFill("solid", fgColor="a8d8a8"),
            'R': PatternFill("solid", fgColor="ffd966"),
            'M': PatternFill("solid", fgColor="f4b183"),
        }
        CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
        THIN   = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Header row
        header = ['Mã HP', 'Tên học phần', 'TC', 'Khối KT'] + plo_mas
        for cidx, h in enumerate(header, 1):
            c = ws.cell(row=1, column=cidx, value=h)
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = CENTER
            c.border = THIN

        # PLO description row (row 2)
        ws.cell(row=2, column=1, value='').border = THIN
        ws.cell(row=2, column=2, value='Mô tả PLO →').font = Font(italic=True, size=8)
        ws.cell(row=2, column=2).border = THIN
        ws.cell(row=2, column=3, value='').border = THIN
        ws.cell(row=2, column=4, value='').border = THIN
        for cidx, plo in enumerate(plos, 5):
            c = ws.cell(row=2, column=cidx, value=(plo.get('mo_ta') or '')[:40])
            c.font = Font(italic=True, size=7, color="444444")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = THIN

        # Data rows
        LEVEL_RANK = {'I': 1, 'R': 2, 'M': 3}
        for ridx, hp_row in enumerate(hp_rows, 3):
            hp_id = hp_row['hp_id']
            clos = [dict(r) for r in self.db.get_clo(hp_id)
                    if not dict(r).get('la_tieu_de_nhom')]

            # Tính mức I/R/M cao nhất cho mỗi PLO
            plo_coverage: Dict[str, str] = {}
            for clo in clos:
                plo = clo.get('cdr_ma', '')
                lvl = clo.get('level_irm', '')
                if plo and lvl:
                    existing_rank = LEVEL_RANK.get(plo_coverage.get(plo, ''), 0)
                    new_rank = LEVEL_RANK.get(lvl, 0)
                    if new_rank > existing_rank:
                        plo_coverage[plo] = lvl

            fill = ALT_FILL if ridx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            row_data = [
                hp_row['ma'],
                hp_row['ten_viet'],
                hp_row['so_tin_chi'],
                hp_row['khoi_kien_thuc'] or '',
            ] + [plo_coverage.get(plo_ma, '') for plo_ma in plo_mas]

            for cidx, val in enumerate(row_data, 1):
                c = ws.cell(row=ridx, column=cidx, value=val)
                c.border = THIN
                if cidx <= 4:
                    c.fill = fill
                    c.alignment = CENTER if cidx in (1, 3) else LEFT
                    c.font = Font(size=9)
                else:
                    plo_ma = plo_mas[cidx - 5]
                    lvl = plo_coverage.get(plo_ma, '')
                    c.fill = ITM_FILL.get(lvl, fill)
                    c.alignment = CENTER
                    c.font = Font(bold=bool(lvl), size=9)

        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 14
        for i in range(5, 5 + len(plo_mas)):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 8

        ws.freeze_panes = 'E3'

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        return True
